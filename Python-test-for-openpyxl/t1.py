# _*_ coding=utf-8 _*_
import os
import openpyxl
'''
filename=input('输入检查文件名：')
wb=openpyxl.load_workbook(filename)
'''
wb=openpyxl.load_workbook('111.xlsx')
for sheetName in wb.get_sheet_names():
    if not sheetName.isdigit():
        continue
'''
sheetname=input('输入检查sheet名：')
sheet1=wb.get_sheet_by_name(sheetname)
'''
sheet1=wb.get_sheet_by_name('Sheet1')
listA=[]
listB=[]
listC=[]
listD=[]
listE=[]
for row in sheet1.iter_rows():
    for cell in row:
        if(str(cell.fill.fgColor.rgb)=='FFFFFF00'):
            str1=str(cell.value)
            print(str(cell.value).find('N/A'))

