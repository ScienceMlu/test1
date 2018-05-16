# _*_ coding=utf-8 _*_

import os
import openpyxl
import t3
from openpyxl.styles import Font, colors, Alignment, fills, PatternFill


from openpyxl import Workbook
wb = Workbook()

ws=wb.create_sheet('Result',index=0)

'''设置字体大小颜色，单元格背景'''


#合并单元格'
ws.merge_cells('A1:B2')
#空值
ws.merge_cells('C1:D2')
#△
ws.merge_cells('E1:F2')
#×
ws.merge_cells('G1:H2')
#N/A
ws.merge_cells('I1:J2')
#格式不符

# 居中单元格
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
ws['I1'].alignment = Alignment(horizontal='center', vertical='center')

# 填写格式
ws['A1'].value = '空值'
ws['C1'].value = '△'
ws['E1'].value = '×'
ws['G1'].value = 'N/A'
ws['I1'].value = '格式不符'
i=0
for cols in range(1,10,2):
    ws.cell(row=3,column=cols,value='个数')
    ws.cell(row=4,column=cols,value=t3.countNumber[i])
    i+=1
i=0
for cols in range(2,12,2):
    ws.cell(row=3,column=cols,value='位置')

#写入位置数据
#空值位置

datalen=len(t3.areaA)+4
print(datalen)
for rows in range(4,datalen):
    ws.cell(column=2,row=rows,value=t3.areaA[i])
    i=i+1
i=0

#△位置

datalen=len(t3.areaB)+4
print(datalen)
for rows in range(4,datalen):
    ws.cell(column=4,row=rows,value=t3.areaB[i])
    i=i+1
i=0

#×位置

datalen=len(t3.areaC)+4
print(datalen)
for rows in range(4,datalen):
    ws.cell(column=6,row=rows,value=t3.areaC[i])
    i=i+1
i=0

#N/A位置

datalen=len(t3.areaD)+4
print(datalen)
for rows in range(4,datalen):
    ws.cell(column=8,row=rows,value=t3.areaD[i])
    i=i+1
i=0

#格式不符合位置

datalen=len(t3.areaE)+4
print(datalen)
for rows in range(4,datalen):
    ws.cell(column=10,row=rows,value=t3.areaE[i])
    i=i+1
i=0

#保存
wb.save(r'Result.xlsx')



