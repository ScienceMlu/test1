# _*_ coding=utf-8 _*_
import os
import openpyxl
import pprint
'''
filename=input('输入检查文件名：')
wb=openpyxl.load_workbook(filename)
'''
wb=openpyxl.load_workbook('111.xlsx')
for sheetName in wb.get_sheet_names():
    if not sheetName.isdigit():
        continue
'''
sheetname=input('输入检查sheet名：')
sheet1=wb.get_sheet_by_name(sheetname)
'''
sheet1=wb.get_sheet_by_name('Sheet1')
listA=[]
listB=[]
listC=[]
listD=[]
listE=[]
areaA=[]
areaB=[]
areaC=[]
areaD=[]
areaE=[]
countNumber=[]
for row in sheet1.iter_rows():
    for cell in row:
        if(str(cell.fill.fgColor.rgb)=='FFFFFF00'):
            if(cell.value==None):
                DataA = [cell.row, cell.column]
                areaA.append(cell.coordinate)
                listA.append(DataA)
                countNumber.append(len(areaA))
                errorA=''
            elif(cell.value=='〇'):
                print('ok')
            elif(str(cell.value).find('×')!= -1):
                DataB = [cell.row, cell.column]
                areaB.append(cell.coordinate)
                listB.append(DataB)
                countNumber.append(len(areaB))
                errorB=''
            elif(cell.value=='△'):
                DataC=[cell.row,cell.column]
                areaC.append(cell.coordinate)
                listC.append(DataC)
                countNumber.append(len(areaC))
                errorC=''
            elif(str(cell.value).find('N/A')!= -1):
                DataD=[cell.row,cell.column]
                listD.append(DataD)
                areaD.append(cell.coordinate)
                countNumber.append(len(areaD))
                errorD=''
            else:
                DataE=[cell.row,cell.column]
                listE.append(DataE)
                areaE.append(cell.coordinate)
                countNumber.append(len(areaE))
                errorE=''


'''                
print('Wait a moment.....')



t=open('coordinate.txt','w')
if(listA!=[]):
    t.write('空值='+pprint.pformat(listA) + '\n' +'个数='+str(len(listA))+'\n'+'====================================='+'\n')
if(listB!=[]):
    t.write('×所在line=' + pprint.pformat(listB)  + '\n' +'个数='+str(len(listB))+ '\n' + '====================================='+'\n')
if(listC!=[]):
    t.write('△所在line=' + pprint.pformat(listC)  + '\n' +'个数='+str(len(listC))+ '\n' + '====================================='+'\n')
if(listD!=[]):
    t.write('N/A所在line='+pprint.pformat(listD)  + '\n' +'个数='+str(len(listD))+ '\n' + '====================================='+'\n')
if(listE!=[]):
    t.write('书写不规范所在line=' + pprint.pformat(listE) + '\n' +'个数='+str(len(listE))+ '\n'+'====================================='+'\n')
t.close()
print('OK,Thank you for using!!!!')
'''